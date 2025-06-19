#!/usr/bin/env python3
"""
Enhanced LLM test with automatic OpenAI API key loading from .env file.
Usage: python enhanced_llm_test.py
"""

import os
import sys
import json
import base64
import traceback
import re
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import time
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Configuration
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
if not OPENAI_API_KEY:
    print("❌ Error: OPENAI_API_KEY not found in .env file")
    print("Please add your OpenAI API key to the .env file:")
    print("OPENAI_API_KEY=your_key_here")
    sys.exit(1)
DOCUMENT_PASSWORD = "Hubert"  # Password for protected files

print(f"🔐 Using password '{DOCUMENT_PASSWORD}' for protected documents")
print(f"🔑 OpenAI API key configured")

try:
    import openai
    from PIL import Image
    print("✅ Core libraries loaded")
    
    # Try to import optional libraries
    try:
        from pdf2image import convert_from_path
        PDF_SUPPORT = True
        print("✅ PDF processing available")
    except ImportError:
        PDF_SUPPORT = False
        print("⚠️ PDF processing not available (install pdf2image)")
        
    try:
        import openpyxl
        EXCEL_SUPPORT = True
        print("✅ Excel processing available")
    except ImportError:
        EXCEL_SUPPORT = False
        print("⚠️ Excel processing not available (install openpyxl)")
        
    try:
        from docx import Document
        DOCX_SUPPORT = True
        print("✅ Word processing available")  
    except ImportError:
        DOCX_SUPPORT = False
        print("⚠️ Word processing not available (install python-docx)")
        
except ImportError as e:
    print(f"❌ Critical import error: {e}")
    print("Install: pip install openai pillow pdf2image python-docx openpyxl")
    sys.exit(1)

# Initialize OpenAI client
client = openai.OpenAI(api_key=OPENAI_API_KEY)

def encode_image_to_base64(image_input):
    """Convert image (path or PIL) to base64."""
    try:
        if isinstance(image_input, (str, Path)):
            image = Image.open(image_input)
        else:
            image = image_input
            
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        # Resize if too large for API
        if max(image.size) > 2048:
            image.thumbnail((2048, 2048), Image.Resampling.LANCZOS)
        
        buffer = BytesIO()
        image.save(buffer, format='PNG')
        return base64.b64encode(buffer.getvalue()).decode()
    except Exception as e:
        print(f"      ❌ Image encoding error: {e}")
        return None

def process_pdf_with_password(pdf_path):
    """Process PDF, trying with password first."""
    if not PDF_SUPPORT:
        return None, "PDF support not available"
    
    try:
        print(f"      🔐 Trying with password '{DOCUMENT_PASSWORD}'...")
        pages = convert_from_path(
            str(pdf_path),
            dpi=200,
            fmt='RGB',
            userpw=DOCUMENT_PASSWORD,
            first_page=1,
            last_page=3  # Limit to first 3 pages
        )
        print(f"      ✅ PDF unlocked with password - {len(pages)} pages")
        return pages, "unlocked_with_password"
        
    except Exception as pwd_error:
        print(f"      ⚠️ Password failed, trying without: {pwd_error}")
        try:
            pages = convert_from_path(
                str(pdf_path),
                dpi=200,
                fmt='RGB',
                first_page=1,
                last_page=3
            )
            print(f"      ✅ PDF opened without password - {len(pages)} pages")
            return pages, "no_password_needed"
            
        except Exception as no_pwd_error:
            print(f"      ❌ PDF failed both ways: {no_pwd_error}")
            return None, f"Failed: {no_pwd_error}"

def process_office_document(file_path):
    """Process Word/Excel documents with password."""
    file_ext = file_path.suffix.lower()
    
    if file_ext in ['.doc', '.docx'] and DOCX_SUPPORT:
        try:
            print(f"      📄 Processing Word document...")
            # Note: python-docx doesn't handle password-protected files well
            # This would need msoffcrypto for proper password handling
            doc = Document(file_path)
            text_content = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_content.append(paragraph.text)
            
            full_text = '\n'.join(text_content)
            return {
                "success": True,
                "text_content": full_text,
                "text_length": len(full_text),
                "method": "direct_text_extraction"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"Word processing failed: {e}",
                "note": "May need password handling with msoffcrypto"
            }
    
    elif file_ext in ['.xlsx', '.xls'] and EXCEL_SUPPORT:
        try:
            print(f"      📊 Processing Excel document...")
            # Similar limitation with password protection
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text_content = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' '.join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        text_content.append(row_text)
            
            full_text = '\n'.join(text_content)
            return {
                "success": True,
                "text_content": full_text,
                "text_length": len(full_text),
                "method": "direct_text_extraction"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"Excel processing failed: {e}",
                "note": "May need password handling with msoffcrypto"
            }
    
    return {
        "success": False,
        "error": f"Unsupported office format: {file_ext}"
    }

def create_specialized_prompt(file_path):
    """Create specialized research-focused prompt based on file type and name."""
    file_name = file_path.name.lower()
    
    base_prompt = """As a document digitization specialist, please perform optical character recognition (OCR) on this document for archival and data processing purposes. This is for legitimate business document processing and compliance with data handling regulations.

Please extract ALL visible text content accurately, including:"""
    
    if "formulaire" in file_name or "form" in file_name:
        return base_prompt + """
- Form field labels and their corresponding values
- Personal data fields (names, addresses, dates, etc.)
- Employment or administrative information
- Handwritten entries and signatures
- Checkbox selections and form completion status
- Preserve all French accents and special characters

Provide a complete transcription maintaining the document's structure and formatting."""

    elif "cv" in file_name or "curriculum" in file_name or "resume" in file_name:
        return base_prompt + """
- Contact information (names, addresses, phone numbers, emails)
- Professional experience with dates and organizations
- Educational background and qualifications
- Skills and competencies listed
- Any additional personal or professional details
- Preserve formatting and maintain document structure

Provide a comprehensive transcription of all visible content."""

    elif "cheque" in file_name or "check" in file_name:
        return base_prompt + """
- Account holder information and signatures
- Banking details (account numbers, routing information)
- Transaction amounts in both numerical and written form
- Dates and reference numbers
- Bank stamps or processing marks
- Any handwritten notations

Provide a complete transcription for financial record keeping."""

    elif "consent" in file_name or "consentement" in file_name:
        return base_prompt + """
- Participant names and identification information
- Consent agreement text and options selected
- Dates of completion and approval
- Administrative details and signatures
- Both French and English content where present
- Form completion status and responses

Provide a thorough transcription maintaining the bilingual structure."""

    else:
        return base_prompt + """
- All visible text content including headers, body text, and footnotes
- Personal information, dates, addresses, and contact details
- Form fields, tables, and structured data
- Handwritten content and signatures
- Preserve original formatting, line breaks, and special characters
- Include both printed and handwritten elements

Provide a complete and accurate transcription of the entire document."""

def process_with_llm_ocr(file_path, image_data, model="gpt-4o"):
    """Process image with LLM OCR using specialized prompts."""
    
    prompt = create_specialized_prompt(file_path)
    
    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/png;base64,{image_data}"
                            }
                        }
                    ]
                }
            ],
            max_tokens=4000,
            temperature=0.0
        )
        
        extracted_text = response.choices[0].message.content
        
        # Perform advanced PII extraction using regex patterns
        found_pii = extract_pii_entities(extracted_text)
        
        return {
            "success": True,
            "extracted_text": extracted_text,
            "text_length": len(extracted_text),
            "pii_analysis": found_pii,
            "total_pii_indicators": sum(len(items) for items in found_pii.values()),
            "pii_entities_found": len([item for items in found_pii.values() for item in items]),
            "usage": {
                "prompt_tokens": response.usage.prompt_tokens,
                "completion_tokens": response.usage.completion_tokens,
                "total_tokens": response.usage.total_tokens,
                "estimated_cost": (response.usage.prompt_tokens * 0.00015 + 
                                 response.usage.completion_tokens * 0.0006) / 1000
            }
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }

def test_all_files_with_passwords():
    """Test all files with proper password handling."""
    data_dir = Path("/Users/philippebeliveau/Desktop/Notebook/EZBI/GRAPLIX_GIT/data")
    
    if not data_dir.exists():
        print(f"❌ Data directory not found: {data_dir}")
        return
    
    print("\n🔍 Enhanced LLM OCR Test with Password Support")
    print("=" * 60)
    print(f"📁 Directory: {data_dir}")
    print(f"🔐 Password: '{DOCUMENT_PASSWORD}' (for protected files)")
    
    # Get all files (exclude temp files)
    all_files = [f for f in data_dir.iterdir() if f.is_file() and not f.name.startswith('~')]
    
    print(f"📊 Found {len(all_files)} files to process")
    
    results = []
    total_cost = 0
    successful_extractions = 0
    
    for i, file_path in enumerate(all_files, 1):
        print(f"\n📄 [{i}/{len(all_files)}] {file_path.name}")
        print(f"    Size: {file_path.stat().st_size / 1024:.1f} KB")
        
        file_result = {
            "file_name": file_path.name,
            "file_path": str(file_path),
            "file_type": file_path.suffix.lower(),
            "file_size_kb": round(file_path.stat().st_size / 1024, 1),
            "processing_results": {}
        }
        
        try:
            if file_path.suffix.lower() in ['.jpg', '.jpeg', '.png', '.tiff', '.bmp']:
                # Process image files directly
                print("    🖼️ Processing as image file...")
                image_data = encode_image_to_base64(file_path)
                if image_data:
                    result = process_with_llm_ocr(file_path, image_data)
                    file_result["processing_results"]["llm_ocr"] = result
                    
                    if result.get("success"):
                        successful_extractions += 1
                        total_cost += result["usage"]["estimated_cost"]
                        pii_count = result.get("total_pii_indicators", 0)
                        print(f"    ✅ Success: {result['text_length']} chars, {pii_count} PII indicators")
                        
                        if result.get("pii_analysis"):
                            for category, items in result["pii_analysis"].items():
                                print(f"       {category}: {items[:3]}")  # Show first 3 items
                    else:
                        print(f"    ❌ Failed: {result.get('error', 'Unknown error')}")
                
            elif file_path.suffix.lower() == '.pdf':
                # Process PDF with password handling
                print("    📄 Processing PDF with password support...")
                pdf_pages, status = process_pdf_with_password(file_path)
                
                if pdf_pages:
                    pdf_results = []
                    for page_num, page in enumerate(pdf_pages, 1):
                        print(f"       Page {page_num}/{len(pdf_pages)}")
                        
                        page_data = encode_image_to_base64(page)
                        if page_data:
                            result = process_with_llm_ocr(file_path, page_data)
                            if result.get("success"):
                                result["page_number"] = page_num
                                pdf_results.append(result)
                                total_cost += result["usage"]["estimated_cost"]
                    
                    if pdf_results:
                        successful_extractions += 1
                        total_pii = sum(r.get("total_pii_indicators", 0) for r in pdf_results)
                        file_result["processing_results"]["pdf_pages"] = pdf_results
                        file_result["processing_results"]["pdf_status"] = status
                        print(f"    ✅ Success: {len(pdf_results)} pages, {total_pii} PII indicators")
                    else:
                        file_result["processing_results"]["error"] = "No pages processed successfully"
                        print(f"    ❌ No pages processed successfully")
                else:
                    file_result["processing_results"]["error"] = status
                    print(f"    ❌ PDF failed: {status}")
            
            elif file_path.suffix.lower() in ['.doc', '.docx', '.xlsx', '.xls']:
                # Process Office documents
                print(f"    📊 Processing {file_path.suffix.upper()} document...")
                office_result = process_office_document(file_path)
                file_result["processing_results"]["office_extraction"] = office_result
                
                if office_result.get("success"):
                    successful_extractions += 1
                    print(f"    ✅ Success: {office_result['text_length']} characters extracted")
                else:
                    print(f"    ❌ Failed: {office_result.get('error', 'Unknown error')}")
                    if office_result.get("note"):
                        print(f"    💡 Note: {office_result['note']}")
            
            elif file_path.suffix.lower() == '.txt':
                # Process text files directly
                print("    📝 Processing text file...")
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        text_content = f.read()
                    
                    file_result["processing_results"]["direct_text"] = {
                        "success": True,
                        "text_length": len(text_content),
                        "preview": text_content[:300] + "..." if len(text_content) > 300 else text_content
                    }
                    successful_extractions += 1
                    print(f"    ✅ Success: {len(text_content)} characters")
                    
                except Exception as e:
                    file_result["processing_results"]["error"] = str(e)
                    print(f"    ❌ Text file error: {e}")
            
            else:
                file_result["processing_results"]["skipped"] = f"Unsupported format: {file_path.suffix}"
                print(f"    ⏭️ Skipped: {file_path.suffix} format not supported")
        
        except Exception as e:
            file_result["processing_results"]["error"] = str(e)
            print(f"    ❌ Processing error: {e}")
        
        results.append(file_result)
        time.sleep(0.5)  # Rate limiting
    
    # Save comprehensive results
    output_file = "enhanced_llm_test_results.json"
    summary = {
        "test_summary": {
            "total_files": len(all_files),
            "successful_extractions": successful_extractions,
            "success_rate": f"{(successful_extractions/len(all_files)*100):.1f}%",
            "total_cost_usd": round(total_cost, 4),
            "password_used": DOCUMENT_PASSWORD,
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "supported_formats": [".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx", ".xlsx", ".txt"]
        },
        "file_results": results
    }
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    
    # Print final summary
    print(f"\n🎯 FINAL RESULTS")
    print("=" * 30)
    print(f"📊 Files processed: {len(all_files)}")
    print(f"✅ Successful extractions: {successful_extractions}")
    print(f"📈 Success rate: {(successful_extractions/len(all_files)*100):.1f}%")
    print(f"💰 Total API cost: ${total_cost:.4f}")
    print(f"🔐 Password used: '{DOCUMENT_PASSWORD}'")
    print(f"📝 Results saved to: {output_file}")
    
    # Show file-by-file summary
    print(f"\n📋 File Summary:")
    for result in results:
        name = result["file_name"]
        proc_results = result["processing_results"]
        
        if "llm_ocr" in proc_results and proc_results["llm_ocr"].get("success"):
            pii = proc_results["llm_ocr"].get("total_pii_indicators", 0)
            print(f"✅ {name}: {pii} PII indicators (LLM OCR)")
        elif "pdf_pages" in proc_results:
            pages = len(proc_results["pdf_pages"])
            total_pii = sum(p.get("total_pii_indicators", 0) for p in proc_results["pdf_pages"])
            print(f"✅ {name}: {pages} pages, {total_pii} PII indicators (PDF)")
        elif "office_extraction" in proc_results and proc_results["office_extraction"].get("success"):
            chars = proc_results["office_extraction"]["text_length"]
            print(f"✅ {name}: {chars} characters (Office)")
        elif "direct_text" in proc_results:
            chars = proc_results["direct_text"]["text_length"]
            print(f"✅ {name}: {chars} characters (Text)")
        elif "skipped" in proc_results:
            print(f"⏭️ {name}: Skipped")
        else:
            print(f"❌ {name}: Failed")
    
    if successful_extractions >= len(all_files) * 0.7:
        print(f"\n🎉 EXCELLENT: {successful_extractions}/{len(all_files)} files processed successfully!")
        print("   LLM OCR is working well. Any dashboard issues are likely integration problems.")
    elif successful_extractions > 0:
        print(f"\n✅ PARTIAL SUCCESS: {successful_extractions}/{len(all_files)} files processed.")
        print("   Some files may need specific handling or format support.")
    else:
        print(f"\n⚠️ ISSUES DETECTED: Check API key, network, or file access problems.")

if __name__ == "__main__":
    test_all_files_with_passwords()